import os
from re import sub
from openpyxl import load_workbook
import pandas as pd

get_all_files = os.listdir("xlsx")

file_paths = [os.path.join("xlsx", file_name) for file_name in get_all_files]
output_paths = [os.path.join("output", file) for file in get_all_files]

for file_path, output_path in zip(file_paths, output_paths):
    all_values = list()
    workbook = load_workbook(
        filename=file_path
    )

    worksheet = workbook['Sheet1']

    all_rows = worksheet.iter_rows(values_only=True)

    first_row = next(all_rows)
    all_values.append(list(first_row))

    values = dict()

    if 'original' in first_row and 'new' in first_row:
        is_word_table = True

        while is_word_table:
            word_row = next(all_rows)
            all_values.append(list(word_row))
            length_of_row = len(word_row)

            if word_row.count(None).__eq__(length_of_row):
                is_word_table = False
            else:
                if word_row.count(None).__eq__(1):
                    none_index = word_row.index(None)
                    for index in range(none_index):
                        values[str(word_row[index])] = str(word_row[none_index + index + 1])
                else:
                    print("something went wrong")

    else:
        raise SyntaxError("original and new is not found in first row.")

    clues_rows = next(all_rows)
    all_values.append(list(clues_rows))
    if 'original clues' in clues_rows and "new clues" in clues_rows:
        new_clues_index = clues_rows.index("new clues")

        for clues in all_rows:
            cut_serial_number = sub(
                pattern=r"^[\d.]+",
                repl="",
                string=clues[0]
            ).strip()

            new_clues = cut_serial_number

            for key, value in zip(values.keys(), values.values()):
                if key in new_clues:
                    new_clues = new_clues.replace(key, value)
                else:
                    pass

            values_data = list(clues)
            values_data[new_clues_index] = new_clues
            all_values.append(values_data)

    else:
        raise SyntaxError("original clues or new clues is not found after the word table")

    workbook.save(file_path)

    df = pd.DataFrame(data=all_values)
    df.to_excel(output_path, index=False, header=False)
