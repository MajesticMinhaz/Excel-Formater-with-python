import os
from re import sub
from dotenv import dotenv_values
from openpyxl import load_workbook


config = dotenv_values(dotenv_path="./config.env")

get_all_files = os.listdir(config.get('FILE_PATH'))

file_paths = [os.path.join(config.get("FILE_PATH"), file_name) for file_name in get_all_files]

for file_path in file_paths:
    workbook = load_workbook(
        filename=file_path
    )

    worksheet = workbook['Sheet1']
    second_worksheet = workbook.create_sheet("result")

    second_worksheet['A1'] = "original"
    second_worksheet['B1'] = "new"

    all_rows = worksheet.iter_rows(values_only=True)

    first_row = next(all_rows)

    values = dict()

    if 'original' in first_row and 'new' in first_row:
        is_word_table = True

        while is_word_table:
            word_row = next(all_rows)
            length_of_row = len(word_row)

            if word_row.count(None).__eq__(length_of_row):
                is_word_table = False
            else:
                if word_row.count(None).__eq__(1):
                    none_index = word_row.index(None)
                    for index in range(none_index):
                        values[str(word_row[index])] = str(word_row[none_index + index + 1])
                else:
                    print("something went wrong")

        for row, (original, new) in enumerate(values.items(), start=2):
            second_worksheet[f"A{row}"] = original
            second_worksheet[f"B{row}"] = new
        second_worksheet.append([''])
    else:
        raise SyntaxError("original and new is not found in first row.")

    original_clues_rows = next(all_rows)
    if 'original clues' in original_clues_rows:
        second_worksheet.append(["original clues", "new clues"])
        for clues in all_rows:
            cut_serial_number = sub(
                pattern=r"^[\d.]+",
                repl="",
                string=clues[0]
            ).strip()

            new_clues = cut_serial_number

            for key, value in zip(values.keys(), values.values()):
                if key in new_clues:
                    new_clues = new_clues.replace(key, value)
                else:
                    pass

            second_worksheet.append([cut_serial_number, new_clues])

    else:
        raise SyntaxError("original clues is not found after the word table")

    workbook.save(file_path)
